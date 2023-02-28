#!/usr/bin/env python
# coding: utf-8

# In[25]:


"""
This code sample shows Custom Model operations with the Azure Form Recognizer client library.
The async versions of the samples require Python 3.6 or later.

To learn more, please visit the documentation - Quickstart: Form Recognizer Python client library SDKs
https://docs.microsoft.com/en-us/azure/applied-ai-services/form-recognizer/quickstarts/try-v3-python-sdk
"""

import pandas as pd
from azure.core.credentials import AzureKeyCredential
from azure.ai.formrecognizer import DocumentAnalysisClient
import re
import os

import numpy as np
from datetime import datetime
import matplotlib.pyplot as plt

"""
Remember to remove the key from your code when you're done, and never post it publicly. For production, use
secure methods to store and access your credentials. For more information, see
https://docs.microsoft.com/en-us/azure/cognitive-services/cognitive-services-security?tabs=command-line%2Ccsharp#environment-variables-and-application-configuration
"""
endpoint = "https://produccion-ml-brazil-south.cognitiveservices.azure.com/"
key = "835f34b1e2ff497ebbb5f4b782a7c682"

model_id = "mov-cuenta-composed-1"


#OTROS LINKS  DE EJEMPLOS
#"https://www.dropbox.com/s/lho0o99no2q9kyv/Test%20Banesco%202%20p%C3%A1g.pdf?dl=0&raw=1"
#'https://www.dropbox.com/s/o9qiumtqa429kus/Bac%20cc%2017%202p.pdf?dl=0&raw=1'
# 'https://www.dropbox.com/s/ov74ny25e9jms1c/Bgeneral_ahorro_3pages.pdf?dl=0&raw=1'
# 'https://www.dropbox.com/s/lho0o99no2q9kyv/Test%20Banesco%202%20p%C3%A1g.pdf?dl=0&raw=1',
#  'https://www.dropbox.com/s/ov74ny25e9jms1c/Bgeneral_ahorro_3pages.pdf?dl=0&raw=1'

data = pd.DataFrame() #data frame vacia para unir (append) todos los archivos


# Replace with the directory containing your local files
local_file_dir = "/Users/abdielgonzalez/Desktop/Panalytics/Data & Analytics/proyectos test/mov-cta/uploads"

# Loop through each file in the directory
for filename in os.listdir(local_file_dir): 
    try: 
        # Only analyze .pdf files
        if filename.endswith(".pdf"):
            # Get the full file path
            file_path = os.path.join(local_file_dir, filename)
            with open(file_path, "rb") as fd:             
                document = fd.read()
            document_analysis_client = DocumentAnalysisClient(
            endpoint=endpoint, credential=AzureKeyCredential(key)
        )    
            poller = document_analysis_client.begin_analyze_document(model=model_id, document=document)
            # Call the begin_analyze_document() method for each file
            result = poller.result().to_dict()

            datos = []   #se crea una lista vacia para luego agregarlo los datos de movimientos de cuenta.
            if 'MONTO' in result['documents'][0]['fields']['movimientos']['value'][0].value.keys():
                for i in result['documents'][0]['fields']['movimientos']['value']:  # loop a traves de cada linea (row)
                    values = i.value['FECHA'].value, i.value['DESCRIPCION'].value, i.value['MONTO'].value, i.value['SALDO'].value  #it creates tuples for each row
                    datos.append(values)  #agrupamos todos los datos en una sola lista de tuples


                column_count = 4 #variable que busca la cantidad de columnas en la tabla, para utilizarlo en las siguientes lineas de codigo
                columns = ['FECHA','DESCRIPCION','MONTO','SALDO']


            else:
                for i in result['documents'][0]['fields']['movimientos']['value']:
                    if 'DEBITO' in i.value.keys():
                        if ((i.value['DEBITO'].value != '0.00') or (i.value['DEBITO'].value != '0') or (i.value['DEBITO'].value != '$0.00')
                        or (i.value['DEBITO'].value != '$0') or (i.value['DEBITO'].value != '$ 0.00') or (i.value['DEBITO'].value != '$ 0')):
                            CREDITO = 0
                            values = i.value['FECHA'].value, i.value['DESCRIPCION'].value, i.value['DEBITO'].value, CREDITO, i.value['SALDO'].value  #it creates tuples for each row
                            datos.append(values)  #agrupamos todos los datos en una sola lista de tuples

                        else:
                            DEBITO = 0
                            values = i.value['FECHA'].value, i.value['DESCRIPCION'].value, DEBITO , i.value['CREDITO'].value, i.value['SALDO'].value  #it creates tuples for each row
                            datos.append(values)  #agrupamos todos los datos en una sola lista de tuples
                    else:
                        DEBITO = 0
                        values = i.value['FECHA'].value, i.value['DESCRIPCION'].value, DEBITO , i.value['CREDITO'].value, i.value['SALDO'].value  #it creates tuples for each row
                        datos.append(values)  #agrupamos todos los datos en una sola lista de tuples

                    column_count = 5 #variable que busca la cantidad de columnas en la tabla, para utilizarlo en las siguientes lineas de codigo
                    columns = ['FECHA','DESCRIPCION','DEBITO','CREDITO','SALDO']


            try:
                datos = list(sum(datos, ()))  #flatten the tuples to list.
            except:
                datos = datos[6:-1]    # OJO CORREGIR este paso excluye la primera fila que por alguna razon no esta con el formato de tuples (), es decir esta sin parentesis, hay que corregir este error.
                datos = list(sum(datos, ()))




            #esta variable toma la lista de todos los datos y los ordena segun la cantidad de columnas, para que pueda ser estructurada como dataframe
            data_group = [datos[i:i+column_count] for i in range(0,len(datos),column_count)]
            #esta variable crea una dataframe a partir de los datos ordenados en columnas
            df = pd.DataFrame(data_group, columns=columns)


            #FUNCION PARA COLOCAR LAS COLUMNAS EN UN ORDEN Y ESCRITURA ESTANDAR
            if len(df.columns) == 5:
                df['DEBITO'] = df['DEBITO'].str.replace(',', '')
                df['DEBITO'] = df['DEBITO'].str.replace('$', '')
                df['DEBITO'] = df['DEBITO'].str.replace(r'\.(?=.*\.)', '', regex=True) #regex = exclude all dots except the last.
                df['CREDITO'] = df['CREDITO'].str.replace(',', '')
                df['CREDITO'] = df['CREDITO'].str.replace('$', '')
                df['CREDITO'] = df['CREDITO'].str.replace(r'\.(?=.*\.)', '', regex=True) #regex = exclude all dots except the last.

                df['SALDO'] = df['SALDO'].str.replace(',', '')
                df['SALDO'] = df['SALDO'].str.replace('$', '')
                df['SALDO'] = df['SALDO'].str.replace(r'\.(?=.*\.)', '', regex=True)  #regex = exclude all dots except the last.

                #transformar los valores de str a numerico (float)
                df['DEBITO'] = pd.to_numeric(df['DEBITO'])
                df['CREDITO'] = pd.to_numeric(df['CREDITO'])
                df['SALDO'] = pd.to_numeric(df['SALDO'])

                # creamos una nueva columna [MONTO], que nos servira para estandarizar el analisis posterior de la DF.
                df = df.fillna(0)                                      #reemplazamos nan values con cero para poder realizar operaciones matematicas
                df['DEBITO'] = df['DEBITO'] * -1                           #transformamos los debitos como transacciones negativas
                df['MONTO'] = df['DEBITO'] + df['CREDITO'] #creamos la nueva columna 'MONTO' (QUE NOS PERMITE ESTANDARIZAR EL PROCESO Y EL CODIGO MAS ADELANTE)
                df = df.drop(columns=['DEBITO', 'CREDITO'])

            elif len(df.columns) == 4:
                df['MONTO'] = df['MONTO'].str.replace(',', '')
                df['MONTO'] = df['MONTO'].str.replace('$', '')
                df['MONTO'] = df['MONTO'].str.replace(r'\.(?=.*\.)', '', regex=True)  #regex = exclude all dots except the last.
                df['SALDO'] = df['SALDO'].str.replace(',', '')
                df['SALDO'] = df['SALDO'].str.replace('$', '')
                df['SALDO'] = df['SALDO'].str.replace(r'\.(?=.*\.)', '', regex=True)  #regex = exclude all dots except the last.

                #transformar los valores de str a numerico (float)
                df['MONTO'] = pd.to_numeric(df['MONTO'])
                df['SALDO'] = pd.to_numeric(df['SALDO'])


            else:
                df.columns =  df.columns
                print('Cantidad de columnas fuera de rango, revisar')

            #SECCION DONDE ESTANDARIZAMOS LA FECHA Y CREAMOS NUEVAS COLUMNAS DE FECHA
            #Creamos una funcion para tomar la fecha, dividirla en dia,mes,año y crear nueva columna con la fecha trasformada en un objeto de fecha (tipo datetime)

            df['FECHA'] = df['FECHA'].str.replace('/|-|" "', "-")
            df['Dia'] = df.FECHA.str.split('-').str[0]
            df['Mes'] = df.FECHA.str.split('-').str[1]
            df['Año'] = df.FECHA.str.split('-').str[2]

                #reemplazamos el formato de meses de texto a numerico
            mes_format = ('1','2','3','4','5','6','7','8','9','10','11','12','01','02','03','04','05','06','07','08','09')
            if all([re.search(w, df['Mes'][1]) for w in mes_format]) is True:

                #unimos la fecha y luego la transformamos en un objeto de datetime, para poder realizar mas funciones.
                df['Fecha'] = df['Dia'] +'/'+ df['Mes'] +'/'+ df['Año']
                df['Fecha'] = pd.to_datetime(df['Fecha'], dayfirst=True)
                df = df.sort_values(by='Fecha')
            else:
                df = df.replace({'Mes': {'ene': '01', 'feb': '02','mar': '03','abr': '04', 'may': '05','jun': '06',
                 'jul': '07', 'ago': '08','sep': '09','oct': '10','nov': '11','dic': '12'}})
                #unimos la fecha y luego la transformamos en un objeto de datetime, para poder realizar mas funciones.
                df['Fecha'] = df['Dia'] +'/'+ df['Mes'] +'/'+ df['Año']
                df['Fecha'] = pd.to_datetime(df['Fecha'], dayfirst=True)
                df = df.sort_values(by='Fecha')


            data = data.append(df, ignore_index=True)
            df = data.drop_duplicates(subset=['FECHA', 'DESCRIPCION', 'MONTO', 'SALDO']) # eliminamos los datos duplicados

            # DETALLE DEL MOVIMIENTO DE CUENTA (banco, # de cuenta y nombre de la cuenta)
            # las mismas provienen directamente de la respuesta de form recognizer.
            detalles_doc = []  #nueva lista para guardar la lista de detalles de cada movimiento de cuenta.

            try:
                if result['documents'][0]['fields']['banco_cuenta']['confidence'] > 0.90:
                    banco_cuenta = result['documents'][0]['fields']['banco_cuenta']['value']
                else:
                    banco_cuenta = None

            except:
                print(nombre_cuenta,'No se logro obtener el nombre del banco')

            try:
                if result['documents'][0]['fields']['numero_cuenta']['confidence'] > 0.90:
                    numero_cuenta = result['documents'][0]['fields']['numero_cuenta']['value']
                else:
                    numero_cuenta = None

            except:
                print('No se logro obtener el numero de cuenta')

            try:
                if result['documents'][0]['fields']['nombre_cuenta']['confidence'] > 0.90:
                    nombre_cuenta = result['documents'][0]['fields']['nombre_cuenta']['value']
                else:
                    nombre_cuenta = None

            except:
                print('No se logro obtener el nombre de cuenta')

            detalles_mov = [banco_cuenta, numero_cuenta, nombre_cuenta]
            detalles_doc.append(detalles_mov)  #se guardan los datos de cada movimiento de cta. en una lista de lista.
            detalles_doc = detalles_doc                                      #******VARIABLE SALIDA API**********
            print(detalles_doc)

            continue

    except:
        print('Links de documentos excluidos por error de lectura:', filename)


#funcion que elimina las comas y signos de dolar, luego transforma los valores a numerico (float)
def clean_currencies(df):
    #eliminar los signos de dollar y comas para trasformar los valores a numerico
    df['MONTO'] = df['MONTO'].str.replace(',', '')
    df['MONTO'] = df['MONTO'].str.replace('$', '')
    df['SALDO'] = df['SALDO'].str.replace(',', '')
    df['SALDO'] = df['SALDO'].str.replace('$', '')

    #transformar los valores de str a numerico (float)
    df['MONTO'] = pd.to_numeric(df['MONTO'])
    df['SALDO'] = pd.to_numeric(df['SALDO'])
    return df

#clean_currencies(df)  #no lo aplico porque ya se aplico mas arriba lo mismo, se deja para referencia y uso alterno de la funcion



################ PERFORMING PREPROCESSING OF INPUT TEXT fron columns 'Descripción'(preparing it for ML Model) ###################
input_column  = 'DESCRIPCION'
def preprocess_text(df,column):
    import re
    for i in range(len(df)):
        ######  REMOVING SPECIAL CHARACTERS
        df.loc[i,column]  = re.sub(r'\W',' ',str(df.loc[i,column]))

        ######  REMOVING ALL SINGLE CHARACTERS
        df.loc[i,column]  = re.sub(r'\s+[a-zA-Z]\s+',' ',str(df.loc[i,column]))

        ######  REMOVING MULTIPLE SPACES WITH SINGLE SPACE
        df.loc[i,column]  = re.sub(r'\s+',' ',str(df.loc[i,column]))

    return df


df = preprocess_text(df,input_column)


#funcion que separa las transacciones en creditos y debitos en columnas separadas
def categorizar_monto(df):
    #creacion de nueva columna para categorizar creditos (c) y debitos (d)
    df['Tipo'] = np.nan
    df['Tipo'] = df['Tipo'].mask(df.MONTO >= 0, 'c').mask(df.MONTO < 0, 'd')
    return df

categorizar_monto(df)
#nuevo data frame de solo creditos
df_creditos = df[df['Tipo']== 'c']
#nuevo data frame de solo debitos
df_debitos = df[df['Tipo']== 'd']
df_debitos.MONTO = df_debitos.MONTO.abs() #transforma los debitos de negativos a numeros absolutos, sin signos.

#funcion para crear nueva columna que categoriza si los creditos es grande en relacion a los demas creditos
def monto_grande(df_creditos,df_debitos):
    df_creditos['Credito_grande'] = np.where(((df_creditos['MONTO']  / df_creditos['MONTO'].sum()) > 0.20), 'Y','N')
    df_debitos['Debito_grande'] = np.where(((df_debitos['MONTO']  / df_debitos['MONTO'].sum()) > 0.20), 'Y','N')

monto_grande(df_creditos,df_debitos)

#Creamos una funcion para predecir la columna Clase, basado en modelo ML(entrenado solo para creditos)
def predecir_clase_creditos(df_creditos):
    #abrimos el modelo de machine learning classifier.pkl' para predecir la columna ['Clase']
    import joblib
    #load from file
    joblib_file = 'classifier_creditos.pkl'
    joblib_model = joblib.load(joblib_file)

    #creamos y predecimos los valores de la nueva columna 'Clase'
    df_creditos['Clase'] = joblib_model.predict(df_creditos['DESCRIPCION'])

predecir_clase_creditos(df_creditos)


#Creamos una funcion para predecir la columna Clase, ahora de df_debitos basado en otro modelo ML (entrenado solo para debitos)
def predecir_clase_debitos(df_debitos):
    #abrimos el modelo de machine learning classifier.pkl' para predecir la columna ['Clase']
    import joblib
    #load from file
    joblib_file = 'classifier_debitos.pkl'
    joblib_model = joblib.load(joblib_file)

    #creamos y predecimos los valores de la nueva columna 'Clase'
    df_debitos['Clase'] = joblib_model.predict(df_debitos['DESCRIPCION'])

predecir_clase_creditos(df_debitos)

#Funcion para determinar si la transaccion es un ingreso o no lo es.
def col_ingreso(df_creditos):
    lst_ingresos = ['ACH ', 'CAJA ', 'BANCA EN LINEA ', 'OTROS ', 'DEPOSITOS ']  #PARA EL DATA SCIENTIST: se debera corregir problema de espacio con la funcion strip (problema se da desde el modelo ML, revisar)
    df_creditos['Ingresos?'] = df_creditos['Clase'].isin(lst_ingresos).map({True: 'Y', False: 'N'}) #aplicamos el cambio de bolean a str

col_ingreso(df_creditos)


from datetime import datetime
#AQUI EMPEZAMOS A CONSTRUIR LAS VARIABLES DE SALIDA DE LA API

df_size = len(df) - 1  #tamaño de la tabla/DataFrame

fecha_inicial = min(df.iloc[df_size,7], df.iloc[0,7])          #******VARIABLE SALIDA API**********
print('fecha inicial:',fecha_inicial.strftime('%d/%m/%Y'))
fecha_final = max(df.iloc[df_size,7], df.iloc[0,7])             #******VARIABLE SALIDA API**********
print('fecha final:',fecha_final.strftime('%d/%m/%Y'))

#determinamos si el orden de las transacciones es ascendente o descendente

if df.iloc[df_size,7] > df.iloc[0,7]:
    ascending_order = True
if ascending_order is True:
    balance_inicial = df.iloc[0,3]     #******VARIABLE SALIDA API**********
    balance_final = df.iloc[df_size,3] #******VARIABLE SALIDA API**********
else:
    print('No se pudo obtener el balance inicial y final')

print('balance_inicial:',balance_inicial)
print('balance_final:', balance_final)



from dateutil.rrule import rrule, MONTHLY
meses_analizados = len([dt for dt in rrule(MONTHLY, dtstart=fecha_inicial, until=fecha_final)])   #******VARIABLE SALIDA API**********
print('meses analizados:',meses_analizados)

total_creditos = df_creditos.MONTO.sum()   #******VARIABLE SALIDA API**********
print('total creditos: ',total_creditos)
total_debitos = df_debitos.MONTO.sum()   #******VARIABLE SALIDA API**********
print('total debitos: ', total_debitos)

deposito_promedio_total = df_creditos.groupby('Mes')['MONTO'].mean().mean()  #******VARIABLE SALIDA API**********
print('Deposito promedio total: ', deposito_promedio_total)
deposito_promedio_min = df_creditos.groupby('Mes')['MONTO'].mean().min()  #******VARIABLE SALIDA API**********
print('Deposito promedio minimo: ', deposito_promedio_min)
deposito_promedio_max = df_creditos.groupby('Mes')['MONTO'].mean().max()   #******VARIABLE SALIDA API**********
print('Deposito promedio maximo: ', deposito_promedio_max)

saldo_promedio_total = df_creditos.groupby('Mes')['SALDO'].mean().mean()   #******VARIABLE SALIDA API**********
print('saldo promedio total: ', saldo_promedio_total)
saldo_promedio_min = df_creditos.groupby('Mes')['SALDO'].mean().min()    #******VARIABLE SALIDA API**********
print('saldo promedio minimo: ', saldo_promedio_min)
saldo_promedio_max = df_creditos.groupby('Mes')['SALDO'].mean().max()    #******VARIABLE SALIDA API**********
print('saldo promedio maximo: ', saldo_promedio_max)

depositos_ingresos = df_creditos[df_creditos['Ingresos?'] == 'Y'].MONTO.sum()   #******VARIABLE SALIDA API**********
print('depositos_ingresos: ', depositos_ingresos)
depositos_no_ingresos = df_creditos[df_creditos['Ingresos?'] == 'N'].MONTO.sum()  #******VARIABLE SALIDA API**********
print('depositos_no_ingresos: ', depositos_no_ingresos)


### EN ESTA SECCION EMPEZAMOS A CONSTRUIR LOS CUADROS QUE SERAN EXPORTADOS A EXCEL

#creamos un cuadro que contenga creditos y debitos (completo) para imprimir luego en el excel
columnsTitles_df = ['Fecha','DESCRIPCION','MONTO','SALDO']
df_xlsx = df.reindex(columns=columnsTitles_df).sort_index() #se reorganiza por row index para que sea tal cual el orden del documento original.

#creamos un cuadro de creditos limpio para exportar a excel
columnsTitles = ['Fecha','DESCRIPCION','MONTO','Clase', 'Credito_grande','Ingresos?']
df_creditos_xlsx = df_creditos.reindex(columns=columnsTitles)

#filtramos una tabla independiente que solo contenga los depositos grandes > 20% de peso del total
df_depositos_grandes_xlsx = df_creditos_xlsx.loc[df_creditos_xlsx['Credito_grande'].isin(['Y'])]

#creamos un cuadro de debitos limpio para exportar a excel
columnsTitles_d = ['Fecha','DESCRIPCION','MONTO','Clase', 'Debito_grande']
df_debitos_xlsx = df_debitos.reindex(columns=columnsTitles_d)

## agrupamos las transacciones por mes, el resultada es un objeto pandas.core.series.
creditos_mensual = df_creditos.groupby('Mes')['MONTO'].mean()
saldos_promedio = df_creditos.groupby('Mes')['SALDO'].mean()
##se transforma el objeto pandas.core.series. a data Frame para poder tratarlo como una tabla
df_creditos_mensual = creditos_mensual.to_frame()
df_saldos_promedio = saldos_promedio.to_frame()

#unimos creditos y saldos promedio en una sola DataFrame
depositos_saldos_mensual_xlsx = pd.DataFrame(df_creditos_mensual)
depositos_saldos_mensual_xlsx['SALDOS_PROMEDIO'] = df_saldos_promedio['SALDO']
depositos_saldos_mensual_xlsx['DEPOSITOS_PROMEDIO'] = depositos_saldos_mensual_xlsx['MONTO']
depositos_saldos_mensual_xlsx.drop('MONTO', axis=1, inplace=True)

#tabla que detalla las transacciones que fueron excluidas de los ingresos
depositos_no_ingresos_xlsx = df_creditos_xlsx[df_creditos_xlsx['Ingresos?'] == 'N']


#AQUI EMPEZAMOS CON LA LIBRERIA PARA ESCRIBIR DATOS EN UN ARCHIVO DE EXCEL
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

wb = Workbook()

ws1 = wb.create_sheet('Data_original',0)
ws2 = wb.create_sheet('Depositos',1)
ws3 = wb.create_sheet('Debitos',2)
ws4 = wb.create_sheet('Promedios_mensuales',3)


#imprimir Data_origilal
for r in dataframe_to_rows(df_xlsx, index=True, header=True):
    ws1.append(r)

for cell in ws1['A'] + ws1[1]:
    cell.style = 'Accent5'

ws1.column_dimensions["A"].width = 4
ws1.column_dimensions["B"].width = 12
ws1.column_dimensions["C"].width = 80
ws1.column_dimensions["D"].width = 14
ws1.column_dimensions["E"].width = 14




########################################################################
#imprimir depositos
offset_row = 8
offset_col = 0

row = 1
for row_data in dataframe_to_rows(df_creditos_xlsx, index=True, header=True):
    col = 1
    for cell_data in row_data:

        ws2.cell(row + offset_row, col + offset_col, cell_data)

        col += 1

    row += 1
for cell in ws2[9]:
    cell.style = 'Accent1'

for cell in ws2['B1:C6'][0] + ws2['B1:C6'][1] + ws2['B1:C6'][2] + ws2['B1:C6'][3] + ws2['B1:C6'][4]+ ws2['B1:C6'][5] + ws2['A1:A8'][7] :
    cell.style = 'Accent1'

ws2.column_dimensions["A"].width = 4
ws2.column_dimensions["B"].width = 16.5
ws2.column_dimensions["C"].width = 80
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 12.5
ws2.column_dimensions["G"].width = 8.5

#escribimos las variables de la API en el encabezado del primer cuadro
ws2['B1'] = 'Fecha inicial: '
ws2['B1'].alignment = Alignment(horizontal='right')
ws2['C1'] = fecha_inicial.strftime('%d/%m/%Y')
ws2['C1'].alignment = Alignment(horizontal='left')

ws2['B2'] = 'Fecha final: '
ws2['B2'].alignment = Alignment(horizontal='right')
ws2['C2'] = fecha_final.strftime('%d/%m/%Y')
ws2['C2'].alignment = Alignment(horizontal='left')

ws2['B3'] = 'Balance inicial: '
ws2['B3'].alignment = Alignment(horizontal='right')
ws2['C3'] = balance_inicial
ws2['C3'].alignment = Alignment(horizontal='left')

ws2['B4'] = 'Balance final: '
ws2['B4'].alignment = Alignment(horizontal='right')
ws2['C4'] = balance_final
ws2['C4'].alignment = Alignment(horizontal='left')

ws2['B5'] = 'Meses analizados: '
ws2['B5'].alignment = Alignment(horizontal='right')
ws2['C5'] = meses_analizados
ws2['C5'].alignment = Alignment(horizontal='left')

ws2['B6'] = 'Total depositos: '
ws2['B6'].alignment = Alignment(horizontal='right')
ws2['C6'] = total_creditos
ws2['C6'].alignment = Alignment(horizontal='left')

#creamos un titulo encabezado para el cuadro
ws2.merge_cells('A8:G8')
cell_titulo = ws2.cell(row=8, column=1)
cell_titulo.value = 'DEPOSITOS TOTALES DEL PERIODO ANALIZADO'
cell_titulo.alignment = Alignment(horizontal='center', vertical='center')

#imprimimos la tabla depositos_no_ingresos_xlsx en la hoja depositos
offset_row = 8
offset_col = 8

row = 1
for row_data in dataframe_to_rows(depositos_no_ingresos_xlsx, index=True, header=True):
    col = 1
    for cell_data in row_data:

        ws2.cell(row + offset_row, col + offset_col, cell_data)

        col += 1

    row += 1
for cell in ws2['J9:O9'][0] + ws2['J1:O8'][7] + ws2['J1:K2'][0] + ws2['J1:K2'][1]:
    cell.style = 'Accent2'


ws2.column_dimensions["I"].width = 4
ws2.column_dimensions["J"].width = 24
ws2.column_dimensions["K"].width = 80
ws2.column_dimensions["L"].width = 14
ws2.column_dimensions["M"].width = 14
ws2.column_dimensions["N"].width = 12.5
ws2.column_dimensions["O"].width = 8.5


#escribimos las variables de la API en el encabezado del primer cuadro
ws2['J1'] = 'Total Depositos NO Ingresos: '
ws2['J1'].alignment = Alignment(horizontal='left')
ws2['K1'] = depositos_no_ingresos
ws2['K1'].alignment = Alignment(horizontal='left')

ws2['J2'] = 'Total Depositos Ingresos: '
ws2['J2'].alignment = Alignment(horizontal='left')
ws2['K2'] = depositos_ingresos
ws2['K2'].alignment = Alignment(horizontal='left')

#creamos un titulo encabezado para el cuadro
ws2.merge_cells('J8:O8')
cell_titulo_2 = ws2.cell(row=8, column=10)
cell_titulo_2.value = 'DETALLE DEPOSITOS NO CONSIDERADOS COMO INGRESOS'
cell_titulo_2.alignment = Alignment(horizontal='center', vertical='center')



##########################################################################
#imprimir debitos
offset_row = 3
offset_col = 0

row = 1
for row_data in dataframe_to_rows(df_debitos_xlsx, index=True, header=True):
    col = 1
    for cell_data in row_data:

        ws3.cell(row + offset_row, col + offset_col, cell_data)

        col += 1

    row += 1
for cell in ws3[4] + ws3['B1:C2'][0]:
    cell.style = 'Accent1'

ws3.column_dimensions["A"].width = 4
ws3.column_dimensions["B"].width = 16.5
ws3.column_dimensions["C"].width = 80
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 12

ws3['B1'] = 'Total debitos: '
ws3['B1'].alignment = Alignment(horizontal='right')
ws3['C1'] = total_debitos
ws3['C1'].alignment = Alignment(horizontal='left')


# DEPOSITOS Y SALDOS PROMEDIOS MENSUALES
#################################################################################
#Escribimos los promedios mensuales de saldos y depositos en la hoja respectiva
for row in dataframe_to_rows(depositos_saldos_mensual_xlsx, index=True, header=True):
    ws4.append(row)

ws4.column_dimensions["A"].width = 4
ws4.column_dimensions["B"].width = 25
ws4.column_dimensions["C"].width = 25
ws4.column_dimensions["D"].width = 25
ws4.column_dimensions["E"].width = 25

#Escribimos los encabezados o resumenes de los valores sobre la tabla
ws4['E16'] = 'Depositos promedio total: '
ws4['E16'].alignment = Alignment(horizontal='right')
ws4['F16'] = deposito_promedio_total

ws4['E17'] = 'Depositos promedio max: '
ws4['E17'].alignment = Alignment(horizontal='right')
ws4['F17'] = deposito_promedio_max

ws4['E18'] = 'Depositos promedio min: '
ws4['E18'].alignment = Alignment(horizontal='right')
ws4['F18'] = deposito_promedio_min

ws4['E19'] = 'Saldos promedio total: '
ws4['E19'].alignment = Alignment(horizontal='right')
ws4['F19'] = saldo_promedio_total

ws4['E20'] = 'Saldos promedio max: '
ws4['E20'].alignment = Alignment(horizontal='right')
ws4['F20'] = saldo_promedio_max

ws4['E21'] = 'Saldos promedio min: '
ws4['E21'].alignment = Alignment(horizontal='right')
ws4['F21'] = saldo_promedio_min

###### ADDING CHART SALDOS Y DEPOSITOS PROMEDIOS ###############
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis

c1 = LineChart()
c1.title = "Saldos y Depósitos Promedio"
c1.style = 2
c1.y_axis.title = 'Monto'
c1.x_axis.title = 'Meses'

data = Reference(ws4, min_col=2, min_row=1, max_col=4, max_row=12)
c1.add_data(data, titles_from_data=True)

ws4.add_chart(c1, "E1")


wb.save("Informe.xlsx")

#link_desgarga_excel   (por hacer???: crear link de descarga, si aplica)    #******VARIABLE SALIDA API**********
